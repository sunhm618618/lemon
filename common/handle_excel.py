import openpyxl


class HandleExcel(object):
    """让读写Excel格式的测试用例更方便

    Attributes:
        filename: Excel文件的路径。
        sheet_name: 要读取的用例在哪个Sheet中，默认是Sheet1
    """
    titles = ('id', 'title', 'data', 'expected')  # 固定的标题

    def __init__(self, filename, sheet_name=None):
        """创建一个HandExcel对象
        args:
            filename: excel文件的路径
            sheet_name: 要读取的工作表名，如果没有指定默认使用Sheet1当做工作表名
        """
        self.filename = filename
        if sheet_name is None:
            self.sheet_name = 'Sheet1'
        else:
            self.sheet_name = sheet_name

    def read_data(self) -> list:  # 返回值是list
        cases = []  # 用来存储所有用例
        titles = []  # 保存标题
        wb = openpyxl.load_workbook(self.filename)  # 获取工作簿对象
        ws = wb[self.sheet_name]

        # 遍历所有行
        # 如果是表头，把表头保存到titles列表中，否则把数据添加到cases列表中
        for i, row in enumerate(ws.rows):
            if i is 0:
                for cell in row:
                    titles.append(cell.value)
            else:
                cases.append(dict(zip(titles, [cell.value for cell in row])))
        return cases

    def write_data(self, row, column, value) -> None:
        """把数据写到指定的单元格

        args:
            row: 行号
            column: 列号
            value:  要写入的值
        returns:
            None
        """
        wb = openpyxl.load_workbook(self.filename)  # 获取一个WorkBook对象
        ws = wb[self.sheet_name]  # 获取一个WorkSheet对象
        ws.cell(row, column, value)
        wb.save(self.filename)

        return None


if __name__ == '__main__':
    excel = HandleExcel(r'C:\Users\10795\PycharmProjects\autotesting\data\注册功能测试用例.xlsx')
    for case in excel.read_data():
        print(case)
